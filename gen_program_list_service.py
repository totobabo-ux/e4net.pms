# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "프로그램목록_Service"

headers = ["시스템명", "프로그램ID", "프로그램명", "클래스명", "클래스경로", "프로그램구분", "개발난이도", "비고"]

header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

BASE_PKG = "com.e4net.pms.service"

# [시스템명, 프로그램ID, 프로그램명, 클래스명, 클래스경로, 프로그램구분, 개발난이도, 비고]
data = [
    (
        "사업관리",
        "DeliverableService",
        "산출물 관리 서비스",
        "DeliverableService",
        BASE_PKG + ".DeliverableService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), save(등록), update(수정), delete(삭제), toDto(DTO 변환), findAllByProject(사업별 전체 조회), upsertFromExcel(엑셀 업로드 upsert)"
    ),
    (
        "사업관리",
        "ManpowerService",
        "투입인력 관리 서비스",
        "ManpowerService",
        BASE_PKG + ".ManpowerService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), save(등록), update(수정), delete(삭제), toDto(DTO 변환)"
    ),
    (
        "사업관리",
        "RequirementService",
        "요구사항 관리 서비스",
        "RequirementService",
        BASE_PKG + ".RequirementService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), findAttachmentById(첨부파일 조회), save(등록+파일업로드), update(수정+파일업로드), delete(삭제), deleteAttachment(첨부 삭제), toDto(DTO 변환), findAllByProject(사업별 전체), upsertFromExcel(엑셀 업로드)"
    ),
    (
        "사업관리",
        "ReqTraceabilityService",
        "요구사항 추적 관리 서비스",
        "ReqTraceabilityService",
        BASE_PKG + ".ReqTraceabilityService",
        "Service", "중",
        "getTabItems(추적 대상 탭 항목 조회), save(추적 연결 저장)"
    ),
    (
        "사업관리",
        "WbsService",
        "WBS 사업일정 관리 서비스",
        "WbsService",
        BASE_PKG + ".WbsService",
        "Service", "중",
        "findByProjectId(사업별 WBS 조회), batchSave(일괄 저장), delete(삭제), createExcelWorkbook(엑셀 생성), upsertFromExcel(엑셀 업로드)"
    ),
    (
        "사업관리",
        "IssueService",
        "이슈 관리 서비스",
        "IssueService",
        BASE_PKG + ".IssueService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), findAttachmentById(첨부 조회), save(등록), update(수정), delete(삭제), deleteAttachment(첨부 삭제), toDto, findAllByProject, createExcelWorkbook(엑셀 생성), upsertFromExcel(엑셀 업로드)"
    ),
    (
        "사업관리",
        "RiskService",
        "위험(리스크) 관리 서비스",
        "RiskService",
        BASE_PKG + ".RiskService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), findAttachmentById(첨부 조회), save(등록), update(수정), delete(삭제), deleteAttachment(첨부 삭제), toDto, findAllByProject, upsertFromExcel(엑셀 업로드), createExcelWorkbook(엑셀 생성)"
    ),
    (
        "사업관리",
        "CustomerReportService",
        "고객보고서 관리 서비스",
        "CustomerReportService",
        BASE_PKG + ".CustomerReportService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), findAttachmentById(첨부 조회), save(등록+파일업로드), update(수정+파일업로드), delete(삭제), deleteAttachment(첨부 삭제), toDto, toAttachFileDto, getAttachmentFilePath / 주간·월간·정기보고·회의록 공용"
    ),
    (
        "사업수행",
        "BusinessFlowService",
        "업무흐름 관리 서비스",
        "BusinessFlowService",
        BASE_PKG + ".BusinessFlowService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), findAttachmentById(첨부 조회), save(등록+파일업로드), update(수정+파일업로드), delete(삭제), deleteAttachment(첨부 삭제), toDto, toAttachFileDto, getAttachmentFilePath"
    ),
    (
        "사업수행",
        "MenuService",
        "사업별 메뉴구조 관리 서비스",
        "MenuService",
        BASE_PKG + ".MenuService",
        "Service", "중",
        "getTreeData(jsTree 트리 데이터 조회), findById(단건 조회), toDto(DTO 변환), findAllByProject(사업별 전체), upsertFromExcel(엑셀 업로드), create(메뉴 추가), update(메뉴 수정), delete(메뉴 삭제)"
    ),
    (
        "사업수행",
        "ScreenListService",
        "화면목록 관리 서비스",
        "ScreenListService",
        BASE_PKG + ".ScreenListService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), findAttachmentById(첨부 조회), save(등록+파일업로드), update(수정), delete(삭제), deleteAttachment(첨부 삭제), toDto, findAllByProject, upsertFromExcel(엑셀 업로드)"
    ),
    (
        "사업수행",
        "ProgramListService",
        "프로그램목록 관리 서비스",
        "ProgramListService",
        BASE_PKG + ".ProgramListService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), findAttachmentById(첨부 조회), save(등록+파일업로드), update(수정), delete(삭제), deleteAttachment(첨부 삭제), toDto, findAllByProject, upsertFromExcel(엑셀 업로드)"
    ),
    (
        "커뮤니티",
        "CommunityService",
        "커뮤니티(공지/자료실) 관리 서비스",
        "CommunityService",
        BASE_PKG + ".CommunityService",
        "Service", "중",
        "search(목록 페이징 조회), findById(단건 조회), findAttachmentById(첨부 조회), save(등록+파일업로드), update(수정), delete(삭제), deleteAttachment(첨부 삭제), toDto, toAttachFileDto, getAttachmentFilePath / 공지사항·자료실 공용"
    ),
    (
        "관리자",
        "ProjectService",
        "사업(프로젝트) 관리 서비스",
        "ProjectService",
        BASE_PKG + ".ProjectService",
        "Service", "중",
        "findAll(전체 목록), search(조건 검색), findById(단건 조회), save(등록), update(수정), delete(삭제), toDto(DTO 변환)"
    ),
    (
        "관리자",
        "CommonCodeService",
        "공통코드 관리 서비스",
        "CommonCodeService",
        BASE_PKG + ".CommonCodeService",
        "Service", "중",
        "getByGroup(그룹별 코드 조회), findAll(전체 조회), findByGroup(그룹 조회), findDistinctGroupCodes(그룹코드 목록), saveNew(신규 등록), update(수정), delete(삭제), upsertFromExcel(엑셀 업로드)"
    ),
    (
        "관리자",
        "AppMenuService",
        "앱 메뉴 관리 서비스",
        "AppMenuService",
        BASE_PKG + ".AppMenuService",
        "Service", "중",
        "getTreeData(jsTree 트리 데이터), findById(단건 조회), toDto(DTO 변환), findAll(전체 목록), create(메뉴 추가), update(메뉴 수정), delete(메뉴 삭제), upsertFromExcel(엑셀 업로드)"
    ),
    (
        "관리자",
        "SystemInfoService",
        "DB 시스템 정보 조회 서비스",
        "SystemInfoService",
        BASE_PKG + ".SystemInfoService",
        "Service", "중",
        "getTables(테이블 목록 조회), getColumns(컬럼 정보 조회), downloadTableList(테이블 목록 엑셀 다운로드), downloadTableDesign(테이블 정의서 엑셀 다운로드)"
    ),
]

category_colors = {
    "사업관리": "EAF4EA",
    "사업수행": "FEF9E7",
    "커뮤니티": "F5EEF8",
    "관리자":   "EBF5FB",
}

col_widths = [12, 26, 28, 26, 46, 12, 10, 80]
for col_idx, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[1].height = 22

for row_no, row_data in enumerate(data, 2):
    sys_name = row_data[0]
    fill_color = category_colors.get(sys_name, "FFFFFF")
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_no, column=col_idx, value=value)
        cell.border = border
        cell.fill = row_fill
        cell.font = Font(size=9)
        if col_idx in (1, 6, 7):
            cell.alignment = center_align
        else:
            cell.alignment = left_align

    ws.row_dimensions[row_no].height = 45

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:" + get_column_letter(len(headers)) + "1"

output_path = r"C:\Users\ilhun\Documents\e4net.pms\e4net_PMS_프로그램목록_Service.xlsx"
wb.save(output_path)
print("저장 완료: " + output_path)
