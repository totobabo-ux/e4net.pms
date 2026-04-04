# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "화면목록"

headers = ["화면ID", "메뉴Level1", "메뉴Level2", "메뉴Level3", "분류", "화면명",
           "화면설명", "URL", "템플릿파일", "비고"]

header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

def screen_id(template_path):
    # "templates/deliverable/list.html" → "deliverable/list"
    path = template_path.replace("templates/", "").replace("\\", "/")
    if path.endswith(".html"):
        path = path[:-5]
    return path

# [화면ID, 메뉴Level1, 메뉴Level2, 메뉴Level3, 분류, 화면명, 화면설명, URL, 템플릿파일, 비고]
data = [
    ("index",                          "공통",    "",         "",             "기타", "로그인",         "사용자 인증 화면. 사번/비밀번호 입력 후 로그인",                            "/",                           "index",                          "비로그인 접근 허용"),
    ("register",                       "공통",    "",         "",             "기타", "회원가입",       "신규 사용자 등록 화면. 사번·이름·소속·연락처 입력",                          "/register",                   "register",                       ""),
    ("project-select",                 "공통",    "",         "",             "기타", "프로젝트 선택",   "로그인 후 참여 중인 사업을 선택하는 화면",                                  "/project-select",             "project-select",                 "로그인 후 필수 진입"),
    ("home",                           "공통",    "",         "",             "기타", "사업 Home",     "선택된 사업의 주요 현황 요약 대시보드",                                     "/home",                       "home",                           ""),
    ("deliverable/list",               "사업관리", "표준관리",   "산출물 관리",   "목록", "산출물 목록",    "사업 산출물 전체 목록 조회. 구분·단계·산출물명 검색",                          "/deliverable",                "deliverable/list",               "엑셀 다운로드·업로드"),
    ("deliverable/form",               "사업관리", "표준관리",   "산출물 관리",   "등록", "산출물 등록",    "신규 산출물 정보(구분·분류·코드·산출물명 등) 입력 및 저장",                     "/deliverable/new",            "deliverable/form",               ""),
    ("deliverable/detail",             "사업관리", "표준관리",   "산출물 관리",   "상세", "산출물 상세",    "산출물 상세 정보 조회. 인라인 수정 전환 가능",                               "/deliverable/{id}",           "deliverable/detail",             ""),
    ("manpower/list",                  "사업관리", "인력관리",   "투입인력 관리", "목록", "인력 목록",     "사업 투입 인력 목록 조회. 역할·급수·소속 검색",                               "/manpower",                   "manpower/list",                  ""),
    ("manpower/form",                  "사업관리", "인력관리",   "투입인력 관리", "등록", "인력 등록",     "투입 인력 정보(이름·역할·급수·투입기간 등) 등록",                             "/manpower/new",               "manpower/form",                  ""),
    ("manpower/detail",                "사업관리", "인력관리",   "투입인력 관리", "상세", "인력 상세",     "투입 인력 상세 정보 조회",                                                 "/manpower/{id}",              "manpower/detail",                ""),
    ("requirement/list",               "사업관리", "범위관리",   "요구사항 관리", "목록", "요구사항 목록",  "사업 요구사항 전체 목록. 유형·우선순위·상태 검색",                             "/requirement",                "requirement/list",               "엑셀 다운로드, 파일 첨부"),
    ("requirement/form",               "사업관리", "범위관리",   "요구사항 관리", "등록", "요구사항 등록",  "요구사항 정보(유형·우선순위·담당자 등) 및 첨부파일 등록",                        "/requirement/new",            "requirement/form",               "다중 파일 첨부"),
    ("requirement/detail",             "사업관리", "범위관리",   "요구사항 관리", "상세", "요구사항 상세",  "요구사항 상세 조회 및 인라인 수정, 첨부파일 다운로드",                           "/requirement/{id}",           "requirement/detail",             "파일 다운로드"),
    ("wbs/list",                       "사업관리", "범위관리",   "사업일정(WBS)", "목록", "WBS 목록/관리", "사업 WBS 트리 구조 조회 및 인라인 편집·일정 관리",                             "/wbs",                        "wbs/list",                       "인라인 편집, 엑셀 다운로드"),
    ("weekly-report/list",             "사업관리", "보고관리",   "주간보고",      "목록", "주간보고 목록",  "주간보고 목록 조회. 보고서명·보고일 검색",                                     "/weekly-report",              "weekly-report/list",             ""),
    ("weekly-report/form",             "사업관리", "보고관리",   "주간보고",      "등록", "주간보고 등록",  "주간보고 정보(보고서명·내용·첨부파일) 등록",                                   "/weekly-report/new",          "weekly-report/form",             "다중 파일 첨부"),
    ("weekly-report/detail",           "사업관리", "보고관리",   "주간보고",      "상세", "주간보고 상세",  "주간보고 상세 조회, 인라인 수정, 첨부파일 다운로드",                             "/weekly-report/{id}",         "weekly-report/detail",           "파일 다운로드"),
    ("monthly-report/list",            "사업관리", "보고관리",   "월간보고",      "목록", "월간보고 목록",  "월간보고 목록 조회. 보고서명·보고일 검색",                                     "/monthly-report",             "monthly-report/list",            ""),
    ("monthly-report/form",            "사업관리", "보고관리",   "월간보고",      "등록", "월간보고 등록",  "월간보고 정보(보고서명·내용·첨부파일) 등록",                                   "/monthly-report/new",         "monthly-report/form",            "다중 파일 첨부"),
    ("monthly-report/detail",          "사업관리", "보고관리",   "월간보고",      "상세", "월간보고 상세",  "월간보고 상세 조회, 인라인 수정, 첨부파일 다운로드",                             "/monthly-report/{id}",        "monthly-report/detail",          "파일 다운로드"),
    ("regular-report/list",            "사업관리", "보고관리",   "정기보고",      "목록", "정기보고 목록",  "정기보고 목록 조회. 보고서명·보고일 검색",                                     "/regular-report",             "regular-report/list",            ""),
    ("regular-report/form",            "사업관리", "보고관리",   "정기보고",      "등록", "정기보고 등록",  "정기보고 정보(보고서명·내용·첨부파일) 등록",                                   "/regular-report/new",         "regular-report/form",            "다중 파일 첨부"),
    ("regular-report/detail",          "사업관리", "보고관리",   "정기보고",      "상세", "정기보고 상세",  "정기보고 상세 조회, 인라인 수정, 첨부파일 다운로드",                             "/regular-report/{id}",        "regular-report/detail",          "파일 다운로드"),
    ("meeting-report/list",            "사업관리", "보고관리",   "회의록",        "목록", "회의록 목록",   "회의록 목록 조회. 회의명·회의일 검색",                                         "/meeting-report",             "meeting-report/list",            ""),
    ("meeting-report/form",            "사업관리", "보고관리",   "회의록",        "등록", "회의록 등록",   "회의록 정보(회의명·내용·참석자·첨부파일) 등록",                                  "/meeting-report/new",         "meeting-report/form",            "다중 파일 첨부"),
    ("meeting-report/detail",          "사업관리", "보고관리",   "회의록",        "상세", "회의록 상세",   "회의록 상세 조회, 인라인 수정, 첨부파일 다운로드",                               "/meeting-report/{id}",        "meeting-report/detail",          "파일 다운로드"),
    ("risk/list",                      "사업관리", "위험관리",   "위험관리",      "목록", "리스크 목록",   "사업 리스크 목록 조회. 유형·심각도·상태 검색",                                   "/risk",                       "risk/list",                      "엑셀 다운로드, 파일 첨부"),
    ("risk/form",                      "사업관리", "위험관리",   "위험관리",      "등록", "리스크 등록",   "리스크 정보(유형·발생확률·영향도·대응방안 등) 등록",                             "/risk/new",                   "risk/form",                      "다중 파일 첨부"),
    ("risk/detail",                    "사업관리", "위험관리",   "위험관리",      "상세", "리스크 상세",   "리스크 상세 조회, 인라인 수정, 첨부파일 다운로드",                               "/risk/{id}",                  "risk/detail",                    "파일 다운로드"),
    ("issue/list",                     "사업관리", "위험관리",   "이슈관리",      "목록", "이슈 목록",    "사업 이슈 목록 조회. 유형·우선순위·상태 검색",                                   "/issue",                      "issue/list",                     "엑셀 다운로드, 파일 첨부"),
    ("issue/form",                     "사업관리", "위험관리",   "이슈관리",      "등록", "이슈 등록",    "이슈 정보(유형·우선순위·담당자·조치내용 등) 등록",                               "/issue/new",                  "issue/form",                     "다중 파일 첨부"),
    ("issue/detail",                   "사업관리", "위험관리",   "이슈관리",      "상세", "이슈 상세",    "이슈 상세 조회, 인라인 수정, 첨부파일 다운로드",                                 "/issue/{id}",                 "issue/detail",                   "파일 다운로드"),
    ("business-flow/list",             "사업수행", "분석",      "업무흐름",      "목록", "업무흐름 목록", "업무흐름 전체 목록 조회",                                                    "/business-flow",              "business-flow/list",             ""),
    ("business-flow/form",             "사업수행", "분석",      "업무흐름",      "등록", "업무흐름 등록", "업무흐름 정보 등록",                                                         "/business-flow/new",          "business-flow/form",             ""),
    ("business-flow/detail",           "사업수행", "분석",      "업무흐름",      "상세", "업무흐름 상세", "업무흐름 상세 조회, 인라인 수정",                                               "/business-flow/{id}",         "business-flow/detail",           ""),
    ("menu-structure/list",            "사업수행", "분석",      "메뉴구조",      "목록", "메뉴구조 관리", "사업별 메뉴 트리 구조 조회 및 CRUD",                                           "/menu-structure",             "menu-structure/list",            "엑셀 업로드"),
    ("screen-list/list",               "사업수행", "설계",      "화면목록",      "목록", "화면목록 목록", "설계 산출물 중 화면목록 전체 조회. 메뉴·분류·화면명 검색",                          "/screen-list",                "screen-list/list",               "엑셀 다운로드·업로드"),
    ("screen-list/form",               "사업수행", "설계",      "화면목록",      "등록", "화면목록 등록", "화면목록 항목(메뉴Level·분류·화면명·URL·템플릿 등) 등록",                          "/screen-list/new",            "screen-list/form",               ""),
    ("screen-list/detail",             "사업수행", "설계",      "화면목록",      "상세", "화면목록 상세", "화면목록 항목 상세 조회, 인라인 수정",                                           "/screen-list/{id}",           "screen-list/detail",             ""),
    ("program-list/list",              "사업수행", "설계",      "프로그램목록",  "목록", "프로그램목록 목록", "프로그램목록 전체 조회. 시스템명·구분·프로그램명 검색",                        "/program-list",               "program-list/list",              "엑셀 다운로드·업로드"),
    ("program-list/form",              "사업수행", "설계",      "프로그램목록",  "등록", "프로그램목록 등록", "프로그램 정보(시스템명·ID·클래스·구분 등) 등록",                             "/program-list/new",           "program-list/form",              ""),
    ("program-list/detail",            "사업수행", "설계",      "프로그램목록",  "상세", "프로그램목록 상세", "프로그램 상세 조회, 인라인 수정",                                            "/program-list/{id}",          "program-list/detail",            ""),
    ("admin/system/database",          "사업수행", "설계",      "테이블목록",    "목록", "DB 테이블 조회", "연결된 DB 테이블 목록 및 컬럼 정보 조회",                                      "/admin/system/database",      "admin/system/database",          "관리자 전용"),
    ("notice/list",                    "커뮤니티", "공지사항",   "공지사항",      "목록", "공지사항 목록", "사업 공지사항 목록 조회. 제목 검색",                                             "/notice",                     "notice/list",                    ""),
    ("notice/form",                    "커뮤니티", "공지사항",   "공지사항",      "등록", "공지사항 등록", "공지사항 제목·내용·첨부파일 등록",                                               "/notice/new",                 "notice/form",                    "다중 파일 첨부"),
    ("notice/detail",                  "커뮤니티", "공지사항",   "공지사항",      "상세", "공지사항 상세", "공지사항 상세 조회, 첨부파일 다운로드",                                           "/notice/{id}",                "notice/detail",                  "파일 다운로드"),
    ("archive/list",                   "커뮤니티", "자료실",    "자료실",        "목록", "자료실 목록",   "자료실 파일 목록 조회. 제목 검색",                                               "/archive",                    "archive/list",                   ""),
    ("archive/form",                   "커뮤니티", "자료실",    "자료실",        "등록", "자료 등록",     "자료 제목·내용·첨부파일 등록",                                                   "/archive/new",                "archive/form",                   "다중 파일 첨부"),
    ("archive/detail",                 "커뮤니티", "자료실",    "자료실",        "상세", "자료 상세",     "자료 상세 조회, 첨부파일 다운로드",                                               "/archive/{id}",               "archive/detail",                 "파일 다운로드"),
    ("project/list",                   "관리자",   "사업 관리",  "사업 목록",    "목록", "사업 목록",    "전체 사업 목록 조회. 사업명·상태 검색",                                           "/projects",                   "project/list",                   ""),
    ("project/form",                   "관리자",   "사업 관리",  "사업 목록",    "등록", "사업 등록",    "신규 사업 정보(사업명·기간·발주처 등) 등록",                                       "/projects/new",               "project/form",                   ""),
    ("admin/user-list",                "관리자",   "사용자 관리", "사용자 목록",  "목록", "사용자 목록",  "전체 사용자 목록 조회 및 관리",                                                   "/admin/users",                "admin/user-list",                "관리자 전용"),
    ("admin/user-form",                "관리자",   "사용자 관리", "사용자 목록",  "수정", "사용자 수정",  "사용자 정보(역할·상태 등) 수정",                                                  "/admin/users/{id}/edit",      "admin/user-form",                "관리자 전용"),
    ("admin/system/common-code",       "관리자",   "시스템 관리", "공통코드 관리", "목록", "공통코드 관리", "공통코드 그룹·코드 CRUD",                                                     "/admin/common-code",          "admin/system/common-code",       "관리자 전용"),
    ("admin/menu/list",                "관리자",   "시스템 관리", "메뉴 관리",    "목록", "메뉴 관리",    "앱 메뉴 트리 구조 조회 및 CRUD",                                                 "/admin/menu",                 "admin/menu/list",                "관리자 전용"),
]

col_widths = [28, 12, 12, 14, 7, 18, 38, 30, 28, 18]
for col_idx, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[1].height = 22

category_colors = {
    "공통":    "EBF5FB",
    "사업관리": "EAF4EA",
    "사업수행": "FEF9E7",
    "커뮤니티": "F5EEF8",
    "관리자":  "FDFEFE",
}

for row_no, row_data in enumerate(data, 2):
    menu1 = row_data[1]
    fill_color = category_colors.get(menu1, "FFFFFF")
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_no, column=col_idx, value=value)
        cell.border = border
        cell.fill = row_fill
        cell.font = Font(size=9)
        if col_idx == 5:
            cell.alignment = center_align
        else:
            cell.alignment = left_align

    ws.row_dimensions[row_no].height = 30

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:" + get_column_letter(len(headers)) + "1"

output_path = r"C:\Users\ilhun\Documents\e4net.pms\e4net_PMS_화면목록.xlsx"
wb.save(output_path)
print("저장 완료: " + output_path)
